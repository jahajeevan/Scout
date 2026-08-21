"""Universal file previewer service.

Given a filesystem path, this module returns everything the Scout frontend
needs to render an inline preview of that file — regardless of format.

Design:
* ``describe(path)`` → metadata + a *kind* tag the UI switches on.
* ``read_text(path)`` → text content for source/markdown/csv/json/txt.
* ``docx_to_html(path)`` → docx rendered as HTML (via mammoth if available).
* ``xlsx_to_json(path)`` → sheets → rows table (via openpyxl if available).
* Binaries the browser handles natively (pdf, images, video, audio) are
  streamed by the FastAPI ``/preview/raw`` endpoint — no conversion needed.

Nothing here writes; nothing here escapes the filesystem outside what the
caller passes in. Path traversal / access control is enforced by the caller.
"""

from __future__ import annotations

import json
import mimetypes
import os
from dataclasses import dataclass, field
from pathlib import Path

_KIND_BY_EXT: dict[str, str] = {
    # browser-native
    "pdf": "pdf",
    "png": "image", "jpg": "image", "jpeg": "image", "gif": "image",
    "webp": "image", "bmp": "image", "tiff": "image", "svg": "image", "ico": "image",
    "mp4": "video", "webm": "video", "mov": "video", "m4v": "video", "mkv": "video",
    "mp3": "audio", "wav": "audio", "ogg": "audio", "m4a": "audio", "flac": "audio", "aac": "audio",
    # text-family
    "txt": "text", "log": "text", "md": "markdown", "markdown": "markdown",
    "csv": "csv", "tsv": "csv", "json": "json", "xml": "text",
    "yaml": "text", "yml": "text", "toml": "text", "ini": "text", "env": "text",
    "html": "html", "htm": "html",
    # code (rendered with syntax highlight by the frontend)
    "py": "code", "js": "code", "ts": "code", "tsx": "code", "jsx": "code",
    "java": "code", "c": "code", "cpp": "code", "h": "code", "hpp": "code",
    "cs": "code", "go": "code", "rs": "code", "rb": "code", "php": "code",
    "swift": "code", "kt": "code", "sh": "code", "bash": "code", "zsh": "code",
    "sql": "code", "css": "code", "scss": "code", "less": "code", "vue": "code",
    # office / structured
    "docx": "docx", "doc": "doc-legacy",
    "xlsx": "xlsx", "xls": "xls-legacy",
    "pptx": "pptx", "ppt": "ppt-legacy",
    # archives
    "zip": "archive", "tar": "archive", "gz": "archive",
    "bz2": "archive", "7z": "archive", "rar": "archive",
}

# Max sizes we will read into memory for text-based previews (heavy PDF/video is streamed).
_MAX_TEXT_BYTES = 2_000_000       # 2 MB
_MAX_XLSX_ROWS = 500              # per sheet, first N rows only
_MAX_DOCX_CHARS = 400_000
_MAX_DIR_ENTRIES = 500


@dataclass
class Preview:
    kind: str                       # 'pdf' | 'image' | 'video' | 'audio' | 'text' | 'code' | 'markdown' | 'html' | 'csv' | 'json' | 'docx' | 'xlsx' | 'directory' | 'archive' | 'unknown'
    name: str
    path: str
    size: int
    mtime: float
    mime: str
    exists: bool = True
    ext: str = ""
    text: str | None = None         # populated for text/code/markdown/json/csv
    html: str | None = None         # populated for docx (mammoth)
    sheets: list[dict] | None = None  # populated for xlsx: [{name, rows, cols, values}]
    entries: list[dict] | None = None  # populated for directory listings
    pages: int | None = None          # populated for pdf: total page count for /preview/pdf-page
    language: str | None = None     # code language hint for the frontend
    truncated: bool = False
    warning: str | None = None      # honest note when we degraded (e.g. lib missing)

    def as_dict(self) -> dict:
        return {
            "kind": self.kind, "name": self.name, "path": self.path,
            "size": self.size, "mtime": self.mtime, "mime": self.mime,
            "exists": self.exists, "ext": self.ext, "text": self.text,
            "html": self.html, "sheets": self.sheets, "entries": self.entries,
            "pages": self.pages,
            "language": self.language, "truncated": self.truncated,
            "warning": self.warning,
        }


def _ext(p: Path) -> str:
    return p.suffix.lstrip(".").lower()


def _kind_for(p: Path) -> str:
    if p.is_dir():
        return "directory"
    return _KIND_BY_EXT.get(_ext(p), "unknown")


def describe(path: str) -> Preview:
    """Cheap metadata-only Preview — no file content read."""
    p = Path(path).expanduser()
    if not p.exists():
        return Preview(
            kind="missing", name=p.name, path=str(p), size=0, mtime=0.0,
            mime="", exists=False,
        )
    stat = p.stat()
    mime = mimetypes.guess_type(str(p))[0] or ("inode/directory" if p.is_dir() else "application/octet-stream")
    kind = _kind_for(p)
    return Preview(
        kind=kind, name=p.name, path=str(p), size=stat.st_size,
        mtime=stat.st_mtime, mime=mime, ext=_ext(p),
    )


def _language_hint(ext: str) -> str | None:
    return {
        "py": "python", "js": "javascript", "ts": "typescript", "tsx": "tsx",
        "jsx": "jsx", "java": "java", "c": "c", "cpp": "cpp", "h": "c",
        "hpp": "cpp", "cs": "csharp", "go": "go", "rs": "rust", "rb": "ruby",
        "php": "php", "swift": "swift", "kt": "kotlin", "sh": "bash",
        "bash": "bash", "zsh": "bash", "sql": "sql", "css": "css",
        "scss": "scss", "less": "less", "vue": "vue", "html": "html",
        "htm": "html", "json": "json", "yaml": "yaml", "yml": "yaml",
        "toml": "toml", "md": "markdown", "markdown": "markdown",
        "csv": "csv", "tsv": "tsv", "xml": "xml",
    }.get(ext)


def _read_text(p: Path, limit: int = _MAX_TEXT_BYTES) -> tuple[str, bool]:
    """Return (text, truncated). Reads at most ``limit`` bytes."""
    with p.open("rb") as fh:
        raw = fh.read(limit + 1)
    truncated = len(raw) > limit
    if truncated:
        raw = raw[:limit]
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="replace")
    return text, truncated


def _preview_directory(p: Path) -> Preview:
    entries: list[dict] = []
    try:
        for i, e in enumerate(sorted(p.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower()))):
            if i >= _MAX_DIR_ENTRIES:
                break
            try:
                st = e.stat()
                entries.append({
                    "name": e.name,
                    "is_dir": e.is_dir(),
                    "size": st.st_size,
                    "mtime": st.st_mtime,
                    "kind": _kind_for(e),
                })
            except OSError:
                continue
    except PermissionError as exc:
        return Preview(
            kind="directory", name=p.name, path=str(p),
            size=0, mtime=p.stat().st_mtime, mime="inode/directory",
            entries=[], warning=f"Permission denied: {exc}",
        )
    return Preview(
        kind="directory", name=p.name or "/", path=str(p),
        size=len(entries), mtime=p.stat().st_mtime, mime="inode/directory",
        entries=entries,
    )


def _preview_docx(p: Path, base: Preview) -> Preview:
    try:
        import mammoth  # type: ignore
    except ImportError:
        base.warning = "Install `mammoth` for rich .docx preview (`pip install mammoth`)."
        return base
    try:
        with p.open("rb") as fh:
            result = mammoth.convert_to_html(fh)
        html = result.value or ""
        truncated = False
        if len(html) > _MAX_DOCX_CHARS:
            html = html[:_MAX_DOCX_CHARS]
            truncated = True
        base.html = html
        base.truncated = truncated
    except Exception as exc:
        base.warning = f"Couldn't render .docx: {exc}"
    return base


def _preview_xlsx(p: Path, base: Preview) -> Preview:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        base.warning = "Install `openpyxl` for .xlsx preview (`pip install openpyxl`)."
        return base
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        sheets: list[dict] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list] = []
            n_rows = 0
            n_cols = 0
            for row in ws.iter_rows(values_only=True):
                if n_rows >= _MAX_XLSX_ROWS:
                    break
                r = [("" if v is None else str(v)) for v in row]
                if r:
                    n_cols = max(n_cols, len(r))
                rows.append(r)
                n_rows += 1
            sheets.append({
                "name": sheet_name, "rows": n_rows, "cols": n_cols, "values": rows,
                "truncated": n_rows >= _MAX_XLSX_ROWS,
            })
        wb.close()
        base.sheets = sheets
        base.truncated = any(s["truncated"] for s in sheets)
    except Exception as exc:
        base.warning = f"Couldn't read .xlsx: {exc}"
    return base


def _preview_pdf_text(p: Path, base: Preview) -> Preview:
    # Populate page count so the frontend can iterate render_pdf_page() calls.
    try:
        import pymupdf  # type: ignore
        doc = pymupdf.open(str(p))
        base.pages = int(doc.page_count)
        doc.close()
    except Exception:
        pass
    try:
        from pypdf import PdfReader
    except ImportError:
        return base
    try:
        reader = PdfReader(str(p))
        text_parts: list[str] = []
        for i, page in enumerate(reader.pages[:20]):
            try:
                text_parts.append(page.extract_text() or "")
            except Exception:
                continue
        text = "\n\n".join(text_parts).strip()
        if text:
            base.text = text[:_MAX_TEXT_BYTES]
            base.truncated = len(reader.pages) > 20 or len(text) > _MAX_TEXT_BYTES
    except Exception as exc:
        base.warning = f"PDF text extraction failed: {exc}"
    return base


def render_pdf_page(path: str, page: int = 0, zoom: float = 1.6) -> tuple[bytes, int]:
    """Render one PDF page as PNG bytes. Returns (png_bytes, total_pages).

    Server-side rendering is required because Electron ships Chromium without
    the PDF plugin, so ``<iframe src=pdf>`` in the FileViewer stays blank.
    """
    import pymupdf  # imported lazily; only needed for PDFs
    doc = pymupdf.open(str(Path(path).expanduser()))
    total = doc.page_count
    idx = max(0, min(int(page), total - 1))
    pg = doc.load_page(idx)
    mat = pymupdf.Matrix(zoom, zoom)
    pix = pg.get_pixmap(matrix=mat, alpha=False)
    data = pix.tobytes("png")
    doc.close()
    return data, total


def build(path: str) -> Preview:
    """Full preview for the given path — content-populated per kind."""
    p = Path(path).expanduser()
    base = describe(str(p))
    if not base.exists:
        return base
    if base.kind == "directory":
        return _preview_directory(p)
    if base.size > 500_000_000:  # 500 MB — refuse in-memory read
        base.warning = f"File is {base.size / 1_000_000:.0f} MB — too large to preview."
        return base

    kind = base.kind
    if kind in ("text", "code", "markdown", "csv", "json", "html"):
        try:
            text, truncated = _read_text(p)
            base.text = text
            base.truncated = truncated
            base.language = _language_hint(base.ext)
        except Exception as exc:
            base.warning = f"Couldn't read as text: {exc}"
    elif kind == "docx":
        base = _preview_docx(p, base)
    elif kind == "xlsx":
        base = _preview_xlsx(p, base)
    elif kind == "pdf":
        base = _preview_pdf_text(p, base)
    # image/video/audio/pdf/archive: metadata only here; the frontend fetches
    # the raw bytes from /preview/raw for the actual rendering.
    return base


def read_bytes(path: str, max_bytes: int | None = None) -> tuple[bytes, str]:
    """Return (bytes, mime) for a file — used by the /preview/raw endpoint."""
    p = Path(path).expanduser()
    if not p.is_file():
        raise FileNotFoundError(path)
    mime = mimetypes.guess_type(str(p))[0] or "application/octet-stream"
    if max_bytes:
        with p.open("rb") as fh:
            return fh.read(max_bytes), mime
    return p.read_bytes(), mime
